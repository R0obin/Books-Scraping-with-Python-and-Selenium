from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import openpyxl

chrome_options = Options()
path = r"C:\Users\Administrator\Desktop\Webscraper/books.xlsx"
wb = openpyxl.load_workbook(path)
sh1 = wb.active
chrome_options.add_experimental_option("detach",True)
url = ("https://books.toscrape.com/index.html")
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
                          options=chrome_options)
chrome_options.binary_location = ".//CFT driver/chrome.exe"
driver.get(url)
page_title = driver.find_element(By.XPATH,"//li[@class='current']").text
page_last_row = page_title.split()[-1]
page_count = int(page_last_row)

for p in range(1,page_count):
    if p == 1:
        Books = driver.find_elements(By.XPATH, "//a[@title]")
        for book in Books:
            book.click()
            driver.implicitly_wait(3)
            Books_title = driver.find_element(By.XPATH, "//h1").text
            Table_row = len(driver.find_elements(By.XPATH,'//table[@class="table table-striped"]/tbody/tr'))
            for r in range(1,Table_row+1):
                Product_info = driver.find_element(By.XPATH,'//*[@class="table table-striped"]/tbody/tr['+str(r)+']/th').text
                Product_values = driver.find_element(By.XPATH,'//*[@class="table table-striped"]/tbody/tr['+str(r)+']/td').text
                Excel_max_row = sh1.max_row
                sh1.cell(row=Excel_max_row+1,column=1).value = Books_title
                sh1.cell(row=Excel_max_row+1, column=2).value = Product_info
                sh1.cell(row=Excel_max_row+1, column=3).value = Product_values
                wb.save(path)
            driver.back()
    else:
        driver.find_element(By.XPATH,"//*[contains(@href,'page-"+str(p)+"')]").click()
        Books = driver.find_elements(By.XPATH, "//a[@title]")
        for book in Books:
            book.click()
            driver.implicitly_wait(4)
            Books_title = driver.find_element(By.XPATH, "//h1").text
            Table_row = len(driver.find_elements(By.XPATH,'//table[@class="table table-striped"]/tbody/tr'))
            for r in range(1,Table_row+1):
                Product_info = driver.find_element(By.XPATH,'//*[@class="table table-striped"]/tbody/tr['+str(r)+']/th').text
                Product_values = driver.find_element(By.XPATH,'//*[@class="table table-striped"]/tbody/tr['+str(r)+']/td').text
                Excel_max_row = sh1.max_row
                sh1.cell(row=Excel_max_row+1,column=1).value = Books_title
                sh1.cell(row=Excel_max_row+1, column=2).value = Product_info
                sh1.cell(row=Excel_max_row+1, column=3).value = Product_values
                wb.save(path)
            driver.back()



